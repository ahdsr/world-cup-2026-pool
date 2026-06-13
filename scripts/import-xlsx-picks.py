from __future__ import annotations

import json
import re
import unicodedata
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path.home() / "Downloads"

WORKBOOK_NAMES = [
    "2026 World Cup Pool - Andrew D (1).xlsx",
    "2026 World Cup Pool - Adam Kolodziej.xlsx",
    "2026 World Cup Pool - Adam Banaszek (2).xlsx",
    "2026 World Cup Pool - Adam Banaszek (1).xlsx",
    "2026 World Cup Pool - Varun.xlsx",
    "2026 World Cup Pool - Tomasz K (2).xlsx",
    "2026 World Cup Pool - Tomasz K (1).xlsx",
    "2026 World Cup Pool - Tadeusz.xlsx",
    "2026 World Cup Pool - Robert.xlsx",
    "2026 World Cup Pool - Rana.xlsx",
    "2026 World Cup Pool - Pawel Jaremko (2).xlsx",
    "2026 World Cup Pool - Pawel Jaremko (1).xlsx",
    "2026 World Cup Pool - Patryk Koscielak.xlsx",
    "2026 World Cup Pool - Michal (2).xlsx",
    "2026 World Cup Pool - Michal (1).xlsx",
    "2026 World Cup Pool - Matthew Wozniczka (2).xlsx",
    "2026 World Cup Pool - Matthew Wozniczka (1).xlsx",
    "2026 World Cup Pool - Matthew Woronko.xlsx",
    "2026 World Cup Pool - Matt C.xlsx",
    "2026 World Cup Pool - Mat Rapacz (2).xlsx",
    "2026 World Cup Pool - Mat Rapacz (1).xlsx",
    "2026 World Cup Pool - Marcin.xlsx",
    "2026 World Cup Pool - Lucas Sokolowski.xlsx",
    "2026 World Cup Pool - Lucas Czuchraj.xlsx",
    "2026 World Cup Pool - Joseph Karam.xlsx",
    "2026 World Cup Pool - Danijel P.xlsx",
    "2026 World Cup Pool - Damon Lim.xlsx",
    "2026 World Cup Pool - Bogdan.xlsx",
    "2026 World Cup Pool - Bianca C.xlsx",
    "2026 World Cup Pool - Andrew D (2).xlsx",
]

SCORING_RULES = {
    "groupAdvancement": 2,
    "exactTopTwoBonus": 3,
    "exactTopFourBonus": 5,
    "roundOf16": 3,
    "quarterFinalists": 5,
    "semifinalists": 7,
    "thirdPlaceMatch": 9,
    "finalists": 10,
    "thirdPlace": 10,
    "runnerUp": 15,
    "champion": 25,
    "bonus": 5,
}

BONUS_ITEMS = [
    ("mostGoalsScored", "Most goals scored", "B2"),
    ("mostGoalsConceded", "Most goals conceded", "B3"),
    ("farthestGoal", "Goal from farthest distance", "B4"),
    ("bestPassCompletion", "Best pass completion %", "B5"),
    ("mostCards", "Most red & yellow cards", "B6"),
]

GROUP_ROWS = {
    "A": 9,
    "B": 16,
    "C": 23,
    "D": 30,
    "E": 37,
    "F": 44,
    "G": 51,
    "H": 58,
    "I": 65,
    "J": 72,
    "K": 79,
    "L": 86,
}

ROUND_OF_32_ROWS = [
    (10, 11),
    (14, 15),
    (18, 19),
    (22, 23),
    (26, 27),
    (30, 31),
    (34, 35),
    (38, 39),
    (44, 45),
    (48, 49),
    (52, 53),
    (56, 57),
    (60, 61),
    (64, 65),
    (68, 69),
    (72, 73),
]
ROUND_OF_16_ROWS = [(12, 13), (20, 21), (28, 29), (36, 37), (46, 47), (54, 55), (62, 63), (70, 71)]
QUARTER_FINAL_ROWS = [(16, 17), (32, 33), (50, 51), (66, 67)]
SEMI_FINAL_ROWS = [(24, 25), (58, 59)]
FINAL_ROWS = (35, 36)
THIRD_PLACE_ROWS = (48, 49)

FLAG_CODES = {
    "Mexico": "mx",
    "South Africa": "za",
    "South Korea": "kr",
    "Czechia": "cz",
    "Canada": "ca",
    "Bosnia & Herzegovina": "ba",
    "Qatar": "qa",
    "Switzerland": "ch",
    "Brazil": "br",
    "Morocco": "ma",
    "Haiti": "ht",
    "Scotland": "gb-sct",
    "United States": "us",
    "Paraguay": "py",
    "Australia": "au",
    "Turkey": "tr",
    "Germany": "de",
    "Curaçao": "cw",
    "Ivory Coast": "ci",
    "Ecuador": "ec",
    "Netherlands": "nl",
    "Japan": "jp",
    "Sweden": "se",
    "Tunisia": "tn",
    "Belgium": "be",
    "Egypt": "eg",
    "Iran": "ir",
    "New Zealand": "nz",
    "Spain": "es",
    "Cape Verde": "cv",
    "Saudi Arabia": "sa",
    "Uruguay": "uy",
    "France": "fr",
    "Senegal": "sn",
    "Iraq": "iq",
    "Norway": "no",
    "Argentina": "ar",
    "Algeria": "dz",
    "Austria": "at",
    "Jordan": "jo",
    "Portugal": "pt",
    "DR Congo": "cd",
    "Uzbekistan": "uz",
    "Colombia": "co",
    "England": "gb-eng",
    "Croatia": "hr",
    "Ghana": "gh",
    "Panama": "pa",
}

CANONICAL_NAMES = {
    "bosnia-herzegovina": "Bosnia & Herzegovina",
    "bosnia and herzegovina": "Bosnia & Herzegovina",
    "curacao": "Curaçao",
    "cote d'ivoire": "Ivory Coast",
    "ivory coast": "Ivory Coast",
    "dr congo": "DR Congo",
    "congo dr": "DR Congo",
    "usa": "United States",
    "united states of america": "United States",
    "turkiye": "Turkey",
}


def clean_text(value: object) -> str:
    return str(value or "").strip()


def normalize(value: object) -> str:
    text = clean_text(value).replace("&", " and ")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def canonical_team(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = text.replace("CuraÃ§ao", "Curaçao")
    return CANONICAL_NAMES.get(normalize(text), text)


def is_selected(value: object) -> bool:
    return clean_text(value).upper() in {"Y", "YES", "1", "TRUE"}


def workbook_owner(path: Path) -> str:
    name = path.stem.removeprefix("2026 World Cup Pool - ").strip()
    return re.sub(r"\s+", " ", name)


def slugify(value: str) -> str:
    text = unicodedata.normalize("NFD", value)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def picks_path_for(owner: str) -> str:
    base = re.sub(r"\s+\(\d+\)$", "", owner)
    if base == "Lucas Sokolowski":
        return "data/picks.json"
    if base == "Marcin":
        return "data/picks-marcin.json"
    if base == "Rana":
        return "data/picks-rana.json"
    return f"data/picks-{slugify(owner)}.json"


def matchup(ws, rows: tuple[int, int], team_col: str, winner_col: str, match_id: str, seeds: list[str] | None = None) -> dict:
    teams = [canonical_team(ws[f"{team_col}{row}"].value) for row in rows]
    winners = [
        team
        for team, row in zip(teams, rows)
        if is_selected(ws[f"{winner_col}{row}"].value)
    ]
    winner = winners[0] if winners else ""
    if not winner and all(teams):
        winner = teams[0]
    if winner not in teams:
        raise ValueError(f"{match_id} winner {winner!r} is not in teams {teams!r}")
    data = {
        "id": match_id,
        "seeds": seeds or [],
        "teams": teams,
        "winner": winner,
    }
    return data


def round_winners(matches: list[dict]) -> list[str]:
    return [match["winner"] for match in matches if match.get("winner")]


def loser(teams: list[str], winner: str) -> str:
    for team in teams:
        if team != winner:
            return team
    return ""


def read_workbook(path: Path, generated_at: str) -> dict:
    workbook = load_workbook(path, data_only=True)
    game = workbook["Game"]

    owner = workbook_owner(path)
    groups = {}
    third_place = {}

    for index, (group_id, start_row) in enumerate(GROUP_ROWS.items()):
        teams = [canonical_team(game[f"D{row}"].value) for row in range(start_row, start_row + 4)]
        predicted_order = [canonical_team(game[f"G{row}"].value) for row in range(start_row, start_row + 4)]
        third_team = predicted_order[2]
        selected = is_selected(game[f"K{9 + index}"].value)
        groups[group_id] = {
            "teams": [{"name": team, "flagCode": FLAG_CODES[team]} for team in teams],
            "predictedOrder": predicted_order,
            "predictedAdvancers": predicted_order[:2] + ([third_team] if selected else []),
        }
        third_place[group_id] = {
            "team": third_team,
            "selected": selected,
        }

    round_of_32 = [
        matchup(
            game,
            rows,
            "O",
            "P",
            f"round-of-32-{index}",
            [clean_text(game[f"N{rows[0]}"].value), clean_text(game[f"N{rows[1]}"].value)],
        )
        for index, rows in enumerate(ROUND_OF_32_ROWS, start=1)
    ]
    round_of_16 = [
        matchup(game, rows, "T", "U", f"round-of-16-{index}")
        for index, rows in enumerate(ROUND_OF_16_ROWS, start=1)
    ]
    quarter_finals = [
        matchup(game, rows, "Y", "Z", f"quarter-final-{index}")
        for index, rows in enumerate(QUARTER_FINAL_ROWS, start=1)
    ]
    semi_finals = [
        matchup(game, rows, "AD", "AE", f"semi-final-{index}")
        for index, rows in enumerate(SEMI_FINAL_ROWS, start=1)
    ]
    final = matchup(game, FINAL_ROWS, "AI", "AJ", "final")
    third_match = matchup(game, THIRD_PLACE_ROWS, "AI", "AJ", "third-place")

    finalists = final["teams"]
    champion = final["winner"]
    runner_up = loser(finalists, champion)

    return {
        "meta": {
            "title": "2026 World Cup Pool Picks",
            "owner": owner,
            "sourceWorkbook": path.name,
            "generatedAt": generated_at,
        },
        "scoringRules": deepcopy(SCORING_RULES),
        "bonus": [
            {
                "id": item_id,
                "label": label,
                "pick": canonical_team(game[cell].value),
            }
            for item_id, label, cell in BONUS_ITEMS
        ],
        "groups": groups,
        "thirdPlace": third_place,
        "knockout": {
            "roundOf32": round_of_32,
            "roundOf16": round_of_16,
            "quarterFinals": quarter_finals,
            "semiFinals": semi_finals,
            "final": {
                "teams": final["teams"],
                "winner": champion,
            },
            "thirdPlace": {
                "teams": third_match["teams"],
                "winner": third_match["winner"],
            },
        },
        "advancement": {
            "roundOf16": round_winners(round_of_32),
            "quarterFinalists": round_winners(round_of_16),
            "semifinalists": round_winners(quarter_finals),
            "finalists": round_winners(semi_finals),
            "thirdPlaceMatch": third_match["teams"],
        },
        "podium": {
            "champion": champion,
            "runnerUp": runner_up,
            "thirdPlace": third_match["winner"],
        },
    }


def validate_picks(picks: dict) -> None:
    owner = picks["meta"]["owner"]
    selected_third = [group for group, value in picks["thirdPlace"].items() if value["selected"]]
    if len(selected_third) != 8:
        raise ValueError(f"{owner}: expected 8 third-place selections, found {len(selected_third)}")

    for group_id, group in picks["groups"].items():
        if len(group["predictedOrder"]) != 4 or any(not team for team in group["predictedOrder"]):
            raise ValueError(f"{owner}: group {group_id} has an incomplete predicted order")
        if len(set(map(normalize, group["predictedOrder"]))) != 4:
            raise ValueError(f"{owner}: group {group_id} has duplicate predicted teams")

    expected_lengths = {
        "roundOf16": 16,
        "quarterFinalists": 8,
        "semifinalists": 4,
        "finalists": 2,
        "thirdPlaceMatch": 2,
    }
    for key, expected in expected_lengths.items():
        actual = picks["advancement"][key]
        if len(actual) != expected or any(not team for team in actual):
            raise ValueError(f"{owner}: {key} expected {expected} teams, found {len(actual)}")

    if not all(picks["podium"].values()):
        raise ValueError(f"{owner}: podium is incomplete")


def entry_sort_key(entry: dict) -> tuple[str, int, str]:
    suffix = re.search(r"\((\d+)\)$", entry["name"])
    base = re.sub(r"\s+\(\d+\)$", "", entry["name"])
    return (base.lower(), int(suffix.group(1)) if suffix else 0, entry["name"].lower())


def main() -> None:
    generated_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    entries_config_path = REPO_ROOT / "data" / "entries.json"
    entries_config = json.loads(entries_config_path.read_text(encoding="utf-8"))
    entries = []

    for workbook_name in WORKBOOK_NAMES:
        path = DOWNLOADS / workbook_name
        if not path.exists():
            raise FileNotFoundError(path)

        picks = read_workbook(path, generated_at)
        validate_picks(picks)

        owner = picks["meta"]["owner"]
        picks_path = picks_path_for(owner)
        (REPO_ROOT / picks_path).write_text(
            json.dumps(picks, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        entries.append(
            {
                "id": slugify(owner),
                "name": owner,
                "picksPath": picks_path,
            }
        )

    entries.sort(key=entry_sort_key)
    entries_config["defaultEntryId"] = "lucas-sokolowski"
    entries_config["entries"] = entries
    entries_config_path.write_text(
        json.dumps(entries_config, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"Imported {len(entries)} entries")


if __name__ == "__main__":
    main()
